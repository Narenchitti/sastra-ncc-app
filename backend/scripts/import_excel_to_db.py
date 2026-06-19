# Ingestion Script: Import Nom-Roll Cadets from Excel to Database
import os
import sys
import uuid
import bcrypt
import logging

# Ensure backend folder is in path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("import_excel")

# Auto-install openpyxl if missing
try:
    import openpyxl
except ImportError:
    logger.info("openpyxl is not installed. Installing it dynamically...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from app.schemas.models import UserBase
from app.services import database

EXCEL_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "Batch 5.xlsx"))

def clean_email(name: str, reg_no: str) -> str:
    """Generates a clean, unique email for each cadet."""
    # Clean name: venkataramanan -> lower case, alphanumeric only
    first_part = "".join(c.lower() for c in name.split()[0] if c.isalnum())
    # If first name is very short, use the next part
    if len(first_part) < 3 and len(name.split()) > 1:
        first_part = "".join(c.lower() for c in name.split()[1] if c.isalnum())
    return f"{first_part}.{reg_no}@sastra.ncc"

async def import_cadets():
    if not os.path.exists(EXCEL_PATH):
        logger.error(f"Excel file not found at: {EXCEL_PATH}")
        return

    logger.info(f"Opening Excel workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb.active

    # Find columns based on headers
    headers = [cell.value for cell in sheet[1]]
    logger.info(f"Detected Headers: {headers}")

    # Expected Header Map
    # Index 1: Reg no, 2: Regtl No, 3: Rank, 4: Name , 5: Date of Birth, 6: Year / Branch, 7: Hostel Info
    col_reg_no = -1
    col_regtl_no = -1
    col_rank = -1
    col_name = -1
    col_dob = -1
    col_branch = -1
    col_hostel = -1

    for idx, h in enumerate(headers):
        if not h:
            continue
        h_str = str(h).strip().lower()
        if "reg no" in h_str:
            col_reg_no = idx
        elif "regtl" in h_str or "regiment" in h_str:
            col_regtl_no = idx
        elif h_str == "rank":
            col_rank = idx
        elif h_str == "name":
            col_name = idx
        elif "birth" in h_str or "dob" in h_str:
            col_dob = idx
        elif "branch" in h_str or "year" in h_str:
            col_branch = idx
        elif "hostel" in h_str or "room" in h_str:
            col_hostel = idx

    if col_reg_no == -1 or col_name == -1 or col_rank == -1:
        logger.error("Required columns (Reg no, Name, Rank) not identified. Ingestion aborted.")
        return

    # Seed default ANO user
    ano_id = "6ced2391-0526-446a-bf3f-32565eb09a0d"
    hashed_ano_pwd = bcrypt.hashpw(b"12345678", bcrypt.gensalt()).decode("utf-8")
    ano_user = UserBase(
        id=ano_id,
        name="Capt. ANO Officer",
        email="ano@sastra.ncc",
        password=hashed_ano_pwd,
        rank="Captain",
        role="ANO",
        batch_year=0,
        regimental_number="ANO/2023/1001",
        registration_number="ANO/1001",
        dob="01-01-1980",
        year_branch="Faculty, Associate NCC Officer",
        hostel_info="Staff Quarters",
        camp_count=0,
        status="APPROVED"
    )
    logger.info("Seeding ANO account (ano@sastra.ncc)...")
    await database.save_user(ano_user)

    count = 0
    # Iterate data rows starting from row 2
    for r_idx in range(2, sheet.max_row + 1):
        row_cells = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, sheet.max_column + 1)]
        
        # Check if row is empty
        if not any(row_cells):
            continue
            
        reg_no = str(sheet.cell(row=r_idx, column=col_reg_no + 1).value or "").strip()
        name = str(sheet.cell(row=r_idx, column=col_name + 1).value or "").strip()
        rank = str(sheet.cell(row=r_idx, column=col_rank + 1).value or "").strip()
        
        if not reg_no or not name:
            continue
            
        regtl_no = str(sheet.cell(row=r_idx, column=col_regtl_no + 1).value or "").strip() if col_regtl_no != -1 else None
        dob = str(sheet.cell(row=r_idx, column=col_dob + 1).value or "").strip() if col_dob != -1 else None
        branch = str(sheet.cell(row=r_idx, column=col_branch + 1).value or "").strip() if col_branch != -1 else None
        hostel = str(sheet.cell(row=r_idx, column=col_hostel + 1).value or "").strip() if col_hostel != -1 else None

        # Clean Rank holder role assignment
        rank_upper = rank.upper()
        role = "cadet"
        if rank_upper == "ANO":
            role = "ANO"

        # Deterministic UUID generation based on Regtl No (stable) or Reg No
        seed_str = regtl_no if regtl_no else reg_no
        uid = str(uuid.uuid5(uuid.NAMESPACE_DNS, seed_str))
        
        # Generate clean email
        email = clean_email(name, reg_no)
        
        # Default password is set to registration number
        hashed_pwd = bcrypt.hashpw(reg_no.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        
        cadet_user = UserBase(
            id=uid,
            name=name,
            email=email,
            password=hashed_pwd,
            rank=rank,
            role=role,
            batch_year=5, # III Year Batch 5
            regimental_number=regtl_no,
            registration_number=reg_no,
            dob=dob,
            year_branch=branch,
            hostel_info=hostel,
            camp_count=0,
            status="APPROVED" # Excel records are pre-approved
        )
        
        logger.info(f"Ingesting: {rank} {name} (Email: {email}, ID: {uid})")
        await database.save_user(cadet_user)
        count += 1

    logger.info(f"SUCCESS: Ingested {count} cadets into the database.")

if __name__ == "__main__":
    import asyncio
    asyncio.run(import_cadets())
